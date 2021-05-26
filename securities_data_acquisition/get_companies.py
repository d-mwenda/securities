from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import select

from db import Company, Category, Session, SecuritiesExchange, Country

p = Path("/home/derick/Downloads").resolve() / "isin-codes_2020_new.xlsx"
w = load_workbook(p, data_only=True)
ws = w["Sheet1"]

session = Session()
companies = dict()
categories = list()

kenya = Country(
    slug="ke",
    name="Kenya"
)
session.add(kenya)

nse = SecuritiesExchange(
    name="Nairobi Securities Exchange",
    slug="NSE",
    country=kenya
)

session.add(nse)

try:
    session.commit()
except:
    session.rollback()
finally:
    session.close()

nse = session.query(SecuritiesExchange).filter(SecuritiesExchange.slug=="NSE").first()

for row in ws.iter_rows(min_row=3, max_col=5, max_row=67):
    categories.append(row[4].value)

categories = set(categories)

for cat in categories:
    c = Category(
        name=cat
    )
    session.add(c)

try:
    session.commit()
except:
    session.rollback()
finally:
    session.close()

for row in ws.iter_rows(min_row=3, max_col=5, max_row=67):
    companies[row[1].value] = Company(
        name=row[0].value,
        ticker_symbol=row[2].value,
        ISIN_CODE=row[1].value,
        securities_exchange=nse,
        shares_issued=row[3].value,
        category=session.query(Category).filter_by(name=row[4].value).first()
    )
    session.add(companies[row[1].value])
    
try:
    session.flush()
    session.commit()
except:
    session.rollback()
finally:
    session.close()
